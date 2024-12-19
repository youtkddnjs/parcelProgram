import openpyxl as op 


testPath = r"/Users/sangwonyoo/VSCODE_Project/Python_Study/"
sampleFile = "priceList.xlsx"

wb = op.load_workbook(testPath + sampleFile)

ws_list = wb.sheetnames

ws01 = wb[ws_list[0]]
ws02 = wb[ws_list[1]]

# ws02의 A열 값과 B열 이후 데이터를 딕셔너리로 저장 (비교를 빠르게 하기 위해)
ws02_data = {}
for row in ws02.iter_rows(min_row=2, max_row=ws02.max_row, min_col=1, max_col=ws02.max_column):
    key = row[0].value  # A열 값
    if key is not None:
        ws02_data[key] = [cell.value for cell in row[1:]]  # B열부터 끝까지 값 저장

# ws01의 A열 값(3번째 행부터 마지막까지)과 ws02 A열 비교
for row_idx, cell in enumerate(ws01['A'][2:], start=3):  # A3부터 시작
    key = cell.value
    if key in ws02_data:  # ws02의 A열에 동일한 값이 있다면
        # ws02에서 가져온 데이터를 ws01의 B열부터 채워 넣기
        for col_idx, value in enumerate(ws02_data[key], start=2):  # B열부터 채움
            ws01.cell(row=row_idx, column=col_idx, value=value)

# 변경된 파일 저장
output_file = f"{testPath}newfile.xlsx"
wb.save(output_file)
print(f"새 파일이 저장되었습니다: {output_file}")