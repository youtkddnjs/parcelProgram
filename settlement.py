import config
import utils
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 색상 정의 함수
def rgb_to_hex(r, g, b):
    return f"{r:02X}{g:02X}{b:02X}"



def createSettlement():
    
    # 테두리 스타일 정의
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 정산하기 시트 생성
    config.new_settle_ws = config.new_wb.create_sheet(title="정산하기")

    # 헤더 셀 설정
    headers = [
        ("A2", "총 발송량", 11, True, 12.83, 60, (241, 169, 131)),
        ("B2", "금액 합계", 11, True, 12.83, 60, (97, 203, 243)),
        ("C2", "택배비", 11, False, 12.83, 60, None),
        ("D2", "반품비", 11, False, 12.83, 60, None),
        ("E2", "제주/도서지역", 11, False, 12.83, 60, None),
        ("F2", "극소", 11, True, 12.83, 60, (181, 230, 162)),
        ("G2", "<--반품", 11, True, 12.83, 60, (181, 230, 162)),
        ("H2", "소", 11, True, 12.83, 60, (181, 230, 162)),
        ("I2", "중", 11, True, 12.83, 60, (181, 230, 162)),
        ("J2", "대1", 11, True, 12.83, 60, (181, 230, 162)),
        ("K2", "대2", 11, True, 12.83, 60, (181, 230, 162)),
        ("L2", "이형", 11, True, 12.83, 60, (181, 230, 162)),
    ]

    # 헤더 작성 및 스타일 적용
    for cell, text, font_size, bold, col_width, row_height, color in headers:
        config.new_settle_ws[cell].value = text
        config.new_settle_ws[cell].font = Font(size=font_size, bold=bold)
        config.new_settle_ws[cell].alignment = Alignment(horizontal="center", vertical="center")
        config.new_settle_ws[cell].border = thin_border  # 테두리 추가
        if color:
            config.new_settle_ws[cell].fill = PatternFill(start_color=rgb_to_hex(*color), end_color=rgb_to_hex(*color), fill_type="solid")
        col_letter = cell[0]
        config.new_settle_ws.column_dimensions[col_letter].width = col_width
        config.new_settle_ws.row_dimensions[2].height = row_height

    # A3 ~ L3 설정
    for col in "ABCDEFGHIJKL":
        config.new_settle_ws.column_dimensions[col].width = 12.83
        cell = f"{col}3"
        config.new_settle_ws[cell].border = thin_border  # 테두리 추가
        config.new_settle_ws[cell].alignment = Alignment(horizontal="center", vertical="center")
    config.new_settle_ws.row_dimensions[3].height = 60

    # 추가 정보 설정
    additional_info = [
        ("B6", "신한은행 : 140-014-171955 / 주식회사 파셀", 20, True, (255, 0, 0), "single"),
        ("B7", "20일 이내 업체명으로 입금 부탁드립니다.", 20, True, (255, 0, 0), "single"),
    ]

    # 추가 정보 작성 및 스타일 적용
    for cell, text, font_size, bold, color, underline in additional_info:
        config.new_settle_ws[cell].value = text
        config.new_settle_ws[cell].font = Font(size=font_size, bold=bold, underline=underline, color=rgb_to_hex(*color))
        config.new_settle_ws[cell].alignment = Alignment(horizontal="left", vertical="center")
        config.new_settle_ws.row_dimensions[int(cell[1])].height = 30


    # 수식 추가
    config.new_settle_ws["A3"].value = "=SUM(F3:L3)"  # F3부터 L3까지의 합산
    config.new_settle_ws["B3"].value = "=SUM(C3:E3)"  # C3부터 E3까지의 합산

    # C3, D3, E3 수식: 다른 워크북/워크시트의 데이터 합산
    title = utils.clean_value(config.new_ws.title)
    config.new_settle_ws["C3"].value = f"=SUM('{title}'!F2:F1048576)"  # new_ws의 6번째 열
    if config.existReturn:
        config.new_settle_ws["D3"].value = f"=SUM('반품'!E2:E1048576)"  # new_return_ws의 5번째 열
    config.new_settle_ws["E3"].value = f"=SUM('{title}'!E2:E1048576)"  # new_ws의 5번째 열


    config.new_settle_ws["F3"].value = config.box_tiny
    config.new_settle_ws["G3"].value = config.returnCount
    config.new_settle_ws["H3"].value = config.box_small  # 소형
    config.new_settle_ws["I3"].value = config.box_medium  # 중형
    config.new_settle_ws["J3"].value = config.box_large1  # 대형1
    config.new_settle_ws["K3"].value = config.box_large2  # 대형2
    config.new_settle_ws["L3"].value = config.box_irregular  # 이형
    config.returnCount = 0 
    config.box_tiny = 0  #극소
    config.box_small = 0  # 소형
    config.box_medium = 0  # 중형
    config.box_large1 = 0  # 대형1
    config.box_large2 = 0  # 대형2
    config.box_irregular = 0  # 이형