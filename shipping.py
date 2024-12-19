import config
import utils
import settlement
import openpyxl as op 


def getcustomerList():
    utils.clear_console()
    if not config.successLoad:
        print("정상적으로 워크북을 로드하지 못했습니다. 작업을 중단합니다.")
        return  # 함수 종료
        
    wb = config.srcWB
    ws_list = wb.sheetnames
    config.srcWS = wb[ws_list[0]] #출고 시트
    config.srcWSreturn =  wb[ws_list[1]] #반품 시트
    
    # A열 데이터 가져오기 (첫 행 제외)
    a_column = [cell.value for cell in config.srcWS['A'][1:] if cell.value is not None]
    # 중복 제거 및 오름차순 정렬
    config.unique_values = sorted(set(a_column), reverse=False) 
    config.companyCount = len(config.unique_values)
    
    print("고객사 리스트:")
    for idx, value in enumerate(config.unique_values, start=1):  # enumerate로 번호 추가, start=1로 시작 값 설정
        print(f"{idx}. {value}")
    
    print(f"업체 수 : {config.companyCount}")
    
def selectCustomer(number):
    utils.clear_console()
    
    if 1 <= number <= len(config.unique_values):  # 범위 확인
        config.target_value = config.unique_values[number - 1]
        print(f"선택된 고객사: {config.target_value}")
        createShipping()
        priceChange()
        createReturn()
        if config.existReturn : 
            priceChangereturn()
        settlement.createSettlement()
        saveFile()
    elif number == 0:
        print(f"전체 고객사 대상")
        
        for idx, value in enumerate(config.unique_values, start=1):  # enumerate를 사용하여 인덱스와 값 가져오기
            config.target_value = value
            print(f"선택된 고객사: {idx}. {config.target_value}")
            createShipping()
            priceChange()
            createReturn()
            if config.existReturn : 
                priceChangereturn()
            settlement.createSettlement()
            saveFile()
        
    else:
        print(f"입력한 번호 {number}가 범위를 벗어났습니다.")
        print(f"1 에서 {len(config.unique_values)} 사이의 값을 입력하세요.")
        

def createShipping():
    
    config.new_wb = op.Workbook()
    config.new_ws = config.new_wb.active
    config.new_ws.title = utils.clean_value(config.target_value)  # 시트 이름을 고객사 이름으로 설정
    
    # 첫 번째 고객사 이름이 있는 모든 행 찾기
    matching_rows = []  # 고객사에 해당하는 모든 행 데이터 저장
    for row in config.srcWS.iter_rows(min_row=2, max_row=config.srcWS.max_row):  # 두 번째 행부터 검사
        if row[0].value == config.target_value:  # A열 값이 target_value와 일치하면
            matching_rows.append([cell.value for cell in row])  # 해당 행 전체 데이터 추가
    
    # 첫 행(헤더) 복사하기
    header = [cell.value for cell in config.srcWS[1]]  # 첫 번째 행(헤더) 가져오기
    config.new_ws.append(header)  # 헤더를 새로운 시트에 추가
    
    # 고객사에 해당하는 모든 행 데이터 추가
    for row in matching_rows:
        config.new_ws.append(row)
        
def createReturn():

    # 첫 번째 고객사 이름이 있는 모든 행 찾기
    matching_rows = []  # 고객사에 해당하는 모든 행 데이터 저장
    for row in config.srcWSreturn.iter_rows(min_row=2, max_row=config.srcWSreturn.max_row):  # 두 번째 행부터 검사
        if row[0].value == config.target_value:  # A열 값이 target_value와 일치하면
            matching_rows.append([cell.value for cell in row])  # 해당 행 전체 데이터 추가
    
    if not matching_rows:
        config.existReturn = False
        print("해당 고객은 반품이 없습니다.")
        return
    
    print("해당 고객은 반품이 있습니다.")
    config.existReturn = True
    config.new_return_ws = config.new_wb.create_sheet(title="반품")
    
    # 첫 행(헤더) 복사하기
    header = [cell.value for cell in config.srcWSreturn[1]]  # 첫 번째 행(헤더) 가져오기
    config.new_return_ws.append(header)  # 헤더를 새로운 시트에 추가
    
    # 고객사에 해당하는 모든 행 데이터 추가
    for row in matching_rows:
        config.new_return_ws.append(row)
    
        
def priceChange():
    wb = config.priceWB
    priceList_ws = wb.active
    
    # 1단계: target_value를 priceListwb의 A열에서 찾기
    found = False
    for cell in priceList_ws['A']:
        if cell.value == config.target_value:
            found = True
            print(f"단가표에서 '{config.target_value}'를 찾았습니다. {cell.row}행")
            config.priceRow = cell.row
            break
        
    if not found:
        print(f"단가표에서 '{config.target_value}'는 없으므로 기본값으로 시작합니다.")
        config.priceRow = 3
        
    row_idx = 2  # 시작 행 (2번째 행, 첫 번째 데이터 행)
    while True:
        # e_value - 무게, f_value - 부피 (두 번째 행, 즉 첫 번째 데이터 행)
        e_value = config.new_ws.cell(row=row_idx, column=5).value  # E열 (5번째 열) 값
        f_value = config.new_ws.cell(row=row_idx, column=6).value  # F열 (6번째 열) 값
        # 값이 없으면 종료
        if e_value is None and f_value is None:
            print("수납 요금 변경 작업 완료")
            config.priceComplete = True
            break

        # priceList_ws에서 e_value와 f_value를 찾기
        priceFound = False
        for col_idx, cell in enumerate(priceList_ws[1], start=1):  # 첫 번째 행 반복
            if cell.value == e_value:  # E열 값과 일치하는 열 찾기
                # 2행의 해당 열에서 F열 값 비교
                if priceList_ws.cell(row=2, column=col_idx).value == f_value:
                    priceFound = True
                    config.priceColumn = col_idx
                    break
                
        if priceFound:
            config.price = priceList_ws.cell(row=config.priceRow, column=config.priceColumn).value
            config.new_ws.cell(row=row_idx, column=8).value = config.price
            # config.priceColumn 값(숫자)을 알파벳으로 변환
            price_column_letter = op.utils.get_column_letter(config.priceColumn)

            # price_column_letter 값에 따라 상자 크기 업데이트
            if price_column_letter in ["C", "D"]:
                config.box_tiny += 1  # 극소
            elif price_column_letter in ["E", "I", "J"]:
                config.box_small += 1  # 소형
            elif price_column_letter in ["F", "K", "N"]:
                config.box_medium += 1  # 중형
            elif price_column_letter == "O":
                config.box_large1 += 1  # 대형1
            elif price_column_letter == "P":
                config.box_large2 += 1  # 대형2
            else:
                config.box_irregular += 1  # 이형
        else:
            config.price = 9999
            config.new_return_ws.cell(row=row_idx, column=7).value = config.price
            print(f"{row_idx}행의 무게 값 '{e_value}'와 부피 값 '{f_value}'를 모두 만족하는 값이 없습니다.")
        row_idx += 1
    
    # E열과 F열 삭제 (E열 = 5번째 열, F열 = 6번째 열)
    config.new_ws.delete_cols(5)  # E열 삭제
    config.new_ws.delete_cols(5)  # F열 삭제 (E열이 삭제되면 F열이 한 칸 앞으로 이동)

def priceChangereturn():
    wb = config.priceWB
    priceList_ws = wb.active
    
    # 1단계: target_value를 priceListwb의 A열에서 찾기
    found = False
    for cell in priceList_ws['A']:
        if cell.value == config.target_value:
            found = True
            print(f"단가표에서 '{config.target_value}'를 찾았습니다. {cell.row}행")
            config.priceRow = cell.row
            break
        
    if not found:
        print(f"단가표에서 '{config.target_value}'는 없으므로 기본값으로 시작합니다.")
        config.priceRow = 3
        
    row_idx = 2  # 시작 행 (2번째 행, 첫 번째 데이터 행)
    while True:
        # 반품값에 따라 금액 재정의
        priceFound = False
        returnPrice = config.new_return_ws.cell(row=row_idx, column=7).value
        # None 값이 아닌 경우에만 처리
        if returnPrice is not None:
            if returnPrice in [1830, 1950]:
                config.priceColumn = 3
                config.box_tiny += 1  # 극소
                priceFound = True
            elif returnPrice == 2400:
                config.priceColumn = 5
                config.box_small += 1  # 소형
                priceFound = True
            elif returnPrice == 3100:
                config.priceColumn = 6
                config.box_medium += 1  # 중형
                priceFound = True
            elif returnPrice > 3100:
                config.priceColumn = 7
                config.box_irregular += 1  # 이형
                priceFound = True
        # 값이 없으면 종료
        if returnPrice is None:
            print("반품 실수령액 작업 완료")
            break
                
        if priceFound:
            config.price = priceList_ws.cell(row=config.priceRow, column=config.priceColumn).value
            config.new_return_ws.cell(row=row_idx, column=7).value = config.price
            config.returnCount += 1
        else:
            config.price = 9999
            config.new_return_ws.cell(row=row_idx, column=7).value = config.price
        
        row_idx += 1
    
    # E열과 F열 삭제 (E열 = 5번째 열, F열 = 6번째 열)
    config.new_return_ws.delete_cols(5)  # E열 삭제
    config.new_return_ws.delete_cols(5)  # F열 삭제 (E열이 삭제되면 F열이 한 칸 앞으로 이동)
    
        
def saveFile():
    # 새로운 파일 저장
    output_file = f"{config.target_value}.xlsx"
    config.new_wb.save(output_file)
    print(f"새 파일이 저장되었습니다: {output_file}")
    config.target_value = ""
    config.new_wb = None
    config.new_ws = None
    config.new_return_ws = None

    